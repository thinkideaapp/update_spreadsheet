import datetime
import numbers
import re

import openpyxl
from openpyxl.styles import Alignment, Font, numbers, NamedStyle
from openpyxl.utils import column_index_from_string

from export_excel import find_last_row_value, read_last_row


def format_date(date):
    # Convert date from "DD/MM/YYYY" to "MARÇO/YYYY"
    date_obj = datetime.datetime.strptime(date, '%d/%m/%Y')
    months = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio',
        'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro',
        'Novembro', 'Dezembro'
    ]
    month_str = months[date_obj.month - 1]
    formatted_date = f"{month_str}/{date_obj.year}"
    return formatted_date


def get_rows_columns(info_dict, bill_group, get_type):
    new_dict = {}
    if bill_group == "A":
        if get_type == "quantity":
            try:
                new_dict = {
                    "AO": info_dict.get("ENERGIA ATIVA FORNECIDA FP", 0),
                    "AP": info_dict.get("ENERGIA ATIVA FORNECIDA HR", 0),
                    "AN": info_dict.get("ENERGIA ATIVA FORNECIDA P", 0),
                    "AX": info_dict.get("ENERGIA INJETADA FP", 0),
                    "AY": info_dict.get("ENERGIA INJETADA HR", 0),
                    "AW": info_dict.get("ENERGIA INJETADA P", 0),
                }
            except:
                pass
        elif get_type == "unit_price":
            try:
                new_dict = {
                    "AC": info_dict.get("ENERGIA ATIVA FORNECIDA FP 1", 0) + info_dict.get("ENERGIA ATIVA FORNECIDA FP 3", 0),
                    "AB": info_dict.get("ENERGIA ATIVA FORNECIDA P 2", 0) + info_dict.get("ENERGIA ATIVA FORNECIDA P 4", 0),
                }
            except:
                pass
        elif get_type == "prices":
            try:
                new_dict = {
                    "Q": info_dict.get("DEMANDA", 0),
                    "S": info_dict.get("UFER FP", 0),
                    "U": info_dict.get("CONTRIB. ILUM. PÚBLICA", 0),
                    "Z": info_dict.get("DEMANDA ULTRAPASSAGEM", 0) + info_dict.get("INDEN. VIOL. PRAZO ATENDIMENTO", 0),
                }
                if new_dict['S'] == 0:
                    new_dict['S'] = info_dict.get("UFER HR", 0)
            except:
                pass
        elif get_type == "kwh_consumed":
            try:
                new_dict = {
                    "AS": info_dict.get("ENERGIA GERAÇÃO 1", 0),
                    "AT": info_dict.get("ENERGIA GERAÇÃO 2", 0),
                    "AU": info_dict.get("ENERGIA GERAÇÃO 3", 0),
                }
            except:
                pass
    else:
        if get_type == "quantity":
            try:
                new_dict = {
                    "AO": info_dict.get("ENERGIA ATIVA FORNECIDA", 0),
                    "AX": info_dict.get("ENERGIA INJETADA", 0),
                }
            except:
                pass
        if get_type == "unit_price":
            try:
                new_dict = {
                    "AB": info_dict.get("ENERGIA ATIVA FORNECIDA 1", 0),
                    "AC": info_dict.get("ENERGIA ATIVA FORNECIDA 1", 0),
                }
            except:
                pass
        if get_type == "prices":
            try:
                new_dict = {
                    "U": info_dict.get("CONTRIB. ILUM. PÚBLICA", 0),
                    "Z": info_dict.get("JUROS MORATÓRIA", 0) + info_dict.get("MULTA", 0),
                }
            except:
                pass
        elif get_type == "kwh_consumed":
            try:
                new_dict = {
                    "AS": info_dict.get("ENERGIA GERAÇÃO 1", 0),
                }
            except:
                pass
    return new_dict


def get_info_rows(text, get_type, bill_group):
    text_init = "ENERGIA ATIVA FORNECIDA"
    text_end = "DEMANDA - kW"
    if get_type == "kwh_consumed":
        text_init = "ENERGIA GERAÇÃO"
    index_init = text.find(text_init)
    index_end = text.find(text_end)
    text_part = text[index_init:index_end]
    rows = text_part.split('\n')

    quantity_rows = [
        "ENERGIA ATIVA FORNECIDA FP",
        "ENERGIA ATIVA FORNECIDA HR",
        "ENERGIA ATIVA FORNECIDA P",
        "ENERGIA INJETADA FP",
        "ENERGIA INJETADA HR",
        "ENERGIA INJETADA P",
    ]
    unit_prices_rows = [
        "ENERGIA ATIVA FORNECIDA FP",
        "ENERGIA ATIVA FORNECIDA P",
        "ENERGIA ATIVA FORNECIDA FP -",
        "ENERGIA ATIVA FORNECIDA P -",
    ]
    prices_rows = [
        "DEMANDA",
        "DEMANDA ULTRAPASSAGEM",
        "UFER FP",
        "CONTRIB. ILUM. PÚBLICA",
        "INDEN. VIOL. PRAZO ATENDIMENTO",
    ]
    kwh_rows = rows[:3]

    if bill_group == "B":
        quantity_rows = [
            "ENERGIA ATIVA FORNECIDA",
            "ENERGIA INJETADA",
        ]
        unit_prices_rows = [
            "ENERGIA ATIVA FORNECIDA",
        ]
        prices_rows = [
            "CONTRIB. ILUM. PÚBLICA",
            "JUROS MORATÓRIA",
            "MULTA",
        ]
        kwh_rows = rows[:1]

    row_list = []
    value = None
    info_dict = {}  # Inicializa o dicionário
    kwh_num = 1
    energy_num = 1
    for i, row in enumerate(rows, start=1):
        if get_type == "quantity":
            if row.split(' kWh')[0] in quantity_rows:
                if bill_group == "A":
                    if len(row.split(' kWh')[0].split(" ")) == 4:
                        value = row.split(' ')[6]
                    else:
                        value = row.split(' ')[5]
                else:
                    if len(row.split(' kWh')[0].split(" ")) == 3:
                        value = row.split(' ')[5]
                    else:
                        value = row.split(' ')[4]
                if value:
                    row_list.append(value)
                    info_dict[row.split(' kWh')[0]] = value
                    quantity_rows.remove(row.split(' kWh')[0])
            elif row.split(' -')[0] in quantity_rows:
                if len(row.split(' -')[0].split(" ")) == 4:
                    value = row.split(' ')[8]
                else:
                    value = row.split(' ')[7]
                row_list.append(value)
                info_dict[row.split(' -')[0]] = value
                quantity_rows.remove(row.split(' -')[0])
        elif get_type == "unit_price":
            if row.split(' kWh')[0] in unit_prices_rows:
                if bill_group == "A":
                    if len(row.split(' kWh')[0].split(" ")) == 4:
                        value = row.split(' ')[5]
                        row_list.append(value)
                    else:
                        value = row.split(' ')[8]
                        row_list.append(value)
                else:
                    if len(row.split(' kWh')[0].split(" ")) == 3:
                        value = row.split(' ')[4]
                        row_list.append(value)
                info_dict[f"{row.split(' kWh')[0]} {energy_num}"] = value
                energy_num += 1
            if row.replace('-', '--').split('- ')[0] in unit_prices_rows:
                if bill_group == "A":
                    if len(row.split(' kWh')[0].split(" ")) == 6:
                        value = row.split(' ')[7]
                        row_list.append(value)
                    else:
                        value = row.split(' ')[8]
                        row_list.append(value)
                else:
                    if len(row.split(' kWh')[0].split(" ")) == 5:
                        value = row.split(' ')[6]
                        row_list.append(value)
                info_dict[f'{row.split(" - ")[0]} {energy_num}'] = value
                energy_num += 1
        elif get_type == "prices":
            if row.split(' kW')[0] in prices_rows:
                if len(row.split(' kW')[0].split(" ")) == 1:
                    value = row.split(' ')[5]
                else:
                    value = row.split(' ')[6]
                row_list.append(value)
                info_dict[row.split(' kW')[0]] = value
                prices_rows.remove(row.split(' kW')[0])
                continue
            if row.split(' kVArh')[0] in ('UFER FP', 'UFER HR'):
                value = row.split(' ')[-1]
                row_list.append(value)
                info_dict[row.split(' kVArh')[0]] = value
                continue
            if row.split(' -')[0] == prices_rows[-1]:
                if bill_group == "A":
                    value = row.split(' ')[-1]
                else:
                    value = re.sub("[a-zA-Z]", "", row.split(" ")[4])
                row_list.append(value)
                info_dict[row.split(' -')[0]] = value
                prices_rows.remove(row.split(' -')[0])
            if row.split(' -')[0] in prices_rows:
                value = re.sub("[a-zA-Z]", "", row.split(" ")[-2])
                if value == '':
                    if row.split(" ")[7] not in ('PONTA', 'FORA PONTA',
                                                 'RESERVADO', 'FORA'):
                        value = re.sub("[a-zA-Z]", "", row.split(" ")[5])
                if value:
                    row_list.append(value)
                    info_dict[row.split(' -')[0]] = value
                    prices_rows.remove(row.split(' -')[0])
            if row.split('.')[0] in prices_rows:
                value = row.split(' ')[-1]
                row_list.append(value)
                info_dict[row.split(' .')[0]] = value
        elif get_type == "kwh_consumed":
            if row in kwh_rows:
                value = row.split(' ')[5]
                row_list.append(value)
                info_dict[f"ENERGIA GERAÇÃO {kwh_num}"] = value
                kwh_num += 1
    try:
        for chave, valor in info_dict.items():
            info_dict[chave] = float(valor.replace('.', '').replace(',', '.'))
        info_dict = get_rows_columns(info_dict, bill_group, get_type)
        return info_dict
    except:
        return []


def bill_classification(text):
    # Expressão regular para encontrar a letra após "Classificação: "
    text_default = r'Classificação:\s*([A-Z])'

    # Procurar o padrão na linha
    match = re.search(text_default, text)

    if match:
        # Retornar a letra encontrada
        return match.group(1)
    else:
        # Caso não encontre o padrão, retornar None ou uma mensagem
        return None


def duplicate_columns_value(ws, last_row, insert_row):
    num_default = re.compile(r'\d+')
    value_a = str(
        ws.cell(row=last_row, column=column_index_from_string("A")).value)
    value_c = str(
        ws.cell(row=last_row, column=column_index_from_string("C")).value)
    value_d = str(
        ws.cell(row=last_row, column=column_index_from_string("D")).value)
    value_e = str(
        ws.cell(row=last_row, column=column_index_from_string("E")).value)
    value_f = str(
        ws.cell(row=last_row, column=column_index_from_string("F")).value)
    value_g = str(
        ws.cell(row=last_row, column=column_index_from_string("G")).value)
    value_h = str(
        ws.cell(row=last_row, column=column_index_from_string("H")).value)
    value_h = num_default.sub(str(insert_row), value_h)
    value_i = str(
        ws.cell(row=last_row, column=column_index_from_string("I")).value)
    value_j = str(
        ws.cell(row=last_row, column=column_index_from_string("J")).value)
    value_k = str(
        ws.cell(row=last_row, column=column_index_from_string("K")).value)
    value_l = str(
        ws.cell(row=last_row, column=column_index_from_string("L")).value)
    value_ag = str(
        ws.cell(row=last_row, column=column_index_from_string("AG")).value)
    value_ag = num_default.sub(str(insert_row), value_ag)
    value_ah = str(
        ws.cell(row=last_row, column=column_index_from_string("AH")).value)
    value_ah = num_default.sub(str(insert_row), value_ah)
    value_ai = str(
        ws.cell(row=last_row, column=column_index_from_string("AI")).value)
    value_ai = num_default.sub(str(insert_row), value_ai)
    value_aj = str(
        ws.cell(row=last_row, column=column_index_from_string("AJ")).value)
    value_aj = num_default.sub(str(insert_row), value_aj)
    value_ak = str(
        ws.cell(row=last_row, column=column_index_from_string("AK")).value)
    value_ak = num_default.sub(str(insert_row), value_ak)
    value_al = str(
        ws.cell(row=last_row, column=column_index_from_string("AL")).value)
    value_al = num_default.sub(str(insert_row), value_al)
    value_am = str(
        ws.cell(row=last_row, column=column_index_from_string("AM")).value)
    value_am = num_default.sub(str(insert_row), value_am)
    value_aq = str(
        ws.cell(row=last_row, column=column_index_from_string("AQ")).value)
    value_aq = num_default.sub(str(insert_row), value_aq)
    value_ar = str(
        ws.cell(row=last_row, column=column_index_from_string("AR")).value)
    value_ar = num_default.sub(str(insert_row), value_ar)
    value_av = str(
        ws.cell(row=last_row, column=column_index_from_string("AV")).value)
    value_av = num_default.sub(str(insert_row), value_av)
    value_az = str(
        ws.cell(row=last_row, column=column_index_from_string("AZ")).value)
    value_az = num_default.sub(str(insert_row), value_az)
    value_ba = str(
        ws.cell(row=last_row, column=column_index_from_string("BA")).value)
    value_ba = num_default.sub(str(insert_row), value_ba)
    value_bb = str(
        ws.cell(row=last_row, column=column_index_from_string("BB")).value)
    value_bb = num_default.sub(str(insert_row), value_bb)
    value_bc = str(
        ws.cell(row=last_row, column=column_index_from_string("BC")).value)
    value_bc = num_default.sub(str(insert_row), value_bc)
    value_bd = str(
        ws.cell(row=last_row, column=column_index_from_string("BD")).value)
    value_bd = num_default.sub(str(insert_row), value_bd)
    value_be = str(
        ws.cell(row=last_row, column=column_index_from_string("BE")).value)
    value_be = num_default.sub(str(insert_row), value_be)
    value_bf = str(
        ws.cell(row=last_row, column=column_index_from_string("BF")).value)
    value_bf = num_default.sub(str(insert_row), value_bf)
    value_bg = str(
        ws.cell(row=last_row, column=column_index_from_string("BG")).value)
    value_bg = num_default.sub(str(insert_row), value_bg)
    value_bh = str(
        ws.cell(row=last_row, column=column_index_from_string("BH")).value)
    value_bh = num_default.sub(str(insert_row), value_bh)
    value_bi = str(
        ws.cell(row=last_row, column=column_index_from_string("BI")).value)
    value_bi = num_default.sub(str(insert_row), value_bi)
    value_bj = str(
        ws.cell(row=last_row, column=column_index_from_string("BJ")).value)
    value_bj = num_default.sub(str(insert_row), value_bj)
    value_bk = str(
        ws.cell(row=last_row, column=column_index_from_string("BK")).value)
    value_bk = num_default.sub(str(insert_row), value_bk)
    value_bl = str(
        ws.cell(row=last_row, column=column_index_from_string("BL")).value)
    value_bl = num_default.sub(str(insert_row), value_bl)
    value_bm = str(
        ws.cell(row=last_row, column=column_index_from_string("BM")).value)
    value_bm = num_default.sub(str(insert_row), value_bm)
    value_bn = str(
        ws.cell(row=last_row, column=column_index_from_string("BN")).value)
    value_bn = num_default.sub(str(insert_row), value_bn)
    value_cn = str(
        ws.cell(row=last_row, column=column_index_from_string("CN")).value)
    value_cn = num_default.sub(str(insert_row), value_cn)
    value_co = str(
        ws.cell(row=last_row, column=column_index_from_string("CO")).value)
    value_co = num_default.sub(str(insert_row), value_co)
    value_cp = str(
        ws.cell(row=last_row, column=column_index_from_string("CP")).value)
    value_cp = num_default.sub(str(insert_row), value_cp)
    value_cq = str(
        ws.cell(row=last_row, column=column_index_from_string("CQ")).value)
    value_cq = num_default.sub(str(insert_row), value_cq)
    value_cr = str(
        ws.cell(row=last_row, column=column_index_from_string("CR")).value)
    value_cr = num_default.sub(str(insert_row), value_cr)

    try:
        value_l = datetime.datetime.strptime(value_l, "%Y-%m-%d %H:%M:%S")
        value_l = value_l.strftime("%d/%m/%Y")
    except:
        pass
    try:
        value_e = datetime.datetime.strptime(value_e, "%Y-%m-%d %H:%M:%S")
        value_e = value_e.strftime("%d/%m/%Y")
    except:
        pass

    ws.cell(row=insert_row, column=column_index_from_string(
        "A")).value = value_a if value_a != "None" else ""
    ws.cell(row=insert_row, column=column_index_from_string(
        "C")).value = value_c if value_c != "None" else ""
    ws.cell(row=insert_row, column=column_index_from_string(
        "D")).value = value_d if value_d != "None" else ""
    ws.cell(row=insert_row, column=column_index_from_string(
        "E")).value = value_e if value_e != "None" else ""
    format_cell = ws.cell(row=insert_row, column=column_index_from_string("F"))
    format_cell.value = value_f.replace('.', ',') if value_f != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    price_cell = ws.cell(row=insert_row, column=column_index_from_string("G"))
    price_cell.value = float(value_g) if value_g != "None" else ""
    price_cell.number_format = 'R$ #,##0.00'
    price_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(row=insert_row, column=column_index_from_string("H"))
    format_cell.value = value_h if value_h != "None" else ""
    format_cell.number_format = 'R$ #,##0.00'
    ws.cell(row=insert_row, column=column_index_from_string(
        "I")).value = value_i if value_i != "None" else ""
    ws.cell(row=insert_row, column=column_index_from_string(
        "J")).value = value_j if value_j != "None" else ""
    ws.cell(row=insert_row, column=column_index_from_string(
        "K")).value = value_k if value_k != "None" else ""
    ws.cell(row=insert_row, column=column_index_from_string(
        "L")).value = value_l if value_l != "None" else ""
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AG"))
    format_cell.value = value_ag if value_ag != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AH"))
    format_cell.value = value_ah if value_ah != "None" else ""
    format_cell.number_format = 'R$ #,##0.00'
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AI"))
    format_cell.value = value_ai if value_ai != "None" else ""
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AJ"))
    format_cell.value = value_aj if value_aj != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AK"))
    format_cell.value = value_ak if value_ak != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AL"))
    format_cell.value = value_al if value_al != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AM"))
    format_cell.value = value_am if value_am != "None" else ""
    format_cell.number_format = 'R$ #,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AQ"))
    format_cell.value = value_aq if value_aq != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AR"))
    format_cell.value = value_ar if value_ar != "None" else ""
    format_cell.number_format = 'R$ #,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AV"))
    format_cell.value = value_av if value_av != "None" else ""
    format_cell.number_format = numbers.FORMAT_NUMBER
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("AZ"))
    format_cell.value = value_az if value_az != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BA"))
    format_cell.value = value_ba if value_ba != "None" else ""
    format_cell.number_format = 'R$ #,##0.00'
    format_cell.alignment = Alignment(horizontal='right')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BB"))
    format_cell.value = value_bb if value_bb != "None" else ""
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BC"))
    format_cell.value = value_bc if value_bc != "None" else ""
    format_cell.number_format = 'R$ #,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BD"))
    format_cell.value = value_bd if value_bd != "None" else ""
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BE"))
    format_cell.value = value_be if value_be != "None" else ""
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BF"))
    format_cell.value = value_bf if value_bf != "None" else ""
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BG"))
    format_cell.value = value_bg if value_bg != "None" else ""
    format_cell.number_format = 'R$ #,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BH"))
    format_cell.value = value_bh if value_bh != "None" else ""
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BI"))
    format_cell.value = value_bi if value_bi != "None" else ""
    format_cell.number_format = 'R$ #,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BJ"))
    format_cell.value = value_bj if value_bj != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BK"))
    format_cell.value = value_bk if value_bk != "None" else ""
    format_cell.number_format = 'R$ #,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BL"))
    format_cell.value = value_bl if value_bl != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BM"))
    format_cell.value = value_bm if value_bm != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("BN"))
    format_cell.value = value_bn if value_bn != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    ws.cell(row=insert_row, column=column_index_from_string(
        "CN")).value = value_cn if value_cn != "None" else ""
    ws.cell(row=insert_row, column=column_index_from_string(
        "CO")).value = value_co if value_co != "None" else ""
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("CP"))
    format_cell.value = value_cp if value_cp != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("CQ"))
    format_cell.value = value_cq if value_cq != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')
    format_cell = ws.cell(
        row=insert_row, column=column_index_from_string("CR"))
    format_cell.value = value_cr if value_cr != "None" else ""
    format_cell.number_format = '#,##0.00'
    format_cell.alignment = Alignment(horizontal='center')


def find_values(text):
    bill_group = bill_classification(text)
    # Definindo padrões de expressões regulares para o valor e a data
    quantity = get_info_rows(text, "quantity", bill_group)
    unit_prices = get_info_rows(text, "unit_price", bill_group)
    prices = get_info_rows(text, "prices", bill_group)
    kwh_consumed = get_info_rows(text, "kwh_consumed", bill_group)

    price_default = r'R\$.*?(\d{1,3}(?:\.\d{3})*,\d{2})'
    date_default = r'(\d{2}/\d{2}/\d{4})'
    interval_default = r'(\d{2}/\d{2}/\d{4})'
    uc_default = r'UC (\d+)'

    # Procurando pelo padrão do valor
    match_valor = re.search(price_default, text)
    match_date = re.findall(date_default, text)
    match_uc = re.search(uc_default, text)
    match_interval = re.findall(interval_default, text)

    if match_valor:
        price = match_valor.group(1)
    if match_date:
        date = match_date[-1]
        date = format_date(date)
    if match_interval:
        cycle = f"{match_interval[2]} a {match_interval[3]}"
    if match_uc:
        uc = match_uc.group(1)

    price = str(price).replace('.', '').replace(',', '.')

    bill_dict = {
        'bill_group': bill_group,
        'price': price,
        'date': date,
        'cycle': cycle,
        'uc': uc,
        'quantity': quantity,
        'unit_price': unit_prices,
        'prices': prices,
        'kwh_consumed': kwh_consumed
    }
    return bill_dict


def organize_sheet_columns(sheet, max_row, bill_dict):
    font_trebuchet_ms = Font(name='Trebuchet MS')

    center_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'M'), value=bill_dict['date'])
    center_cell.alignment = Alignment(horizontal='center')
    center_cell.font = font_trebuchet_ms

    center_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'N'), value=bill_dict['cycle'])
    center_cell.alignment = Alignment(horizontal='center')
    center_cell.font = font_trebuchet_ms

    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'O'), value=float(bill_dict['price']))
    price_cell.number_format = 'R$ #,##0.00'
    price_cell.font = font_trebuchet_ms
    price_cell.alignment = Alignment(horizontal='right')

    center_cell = sheet.cell(
        row=max_row, column=column_index_from_string('B'), value=int(
            bill_dict['uc']))
    center_cell.alignment = Alignment(horizontal='center')
    center_cell.font = font_trebuchet_ms

    if bill_dict['bill_group'] == 'A':
        if 'AO' in bill_dict['quantity']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AO'), value=float(bill_dict['quantity']['AO']))
            price_cell.font = font_trebuchet_ms
            price_cell.number_format = '#,##0.00'
            price_cell.alignment = Alignment(horizontal='center')

        if 'AP' in bill_dict['quantity']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AP'), value=float(bill_dict['quantity']['AP']))
            price_cell.font = font_trebuchet_ms

        if 'AN' in bill_dict['quantity']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AN'), value=float(bill_dict['quantity']['AN']))
            price_cell.font = font_trebuchet_ms

        if 'AX' in bill_dict['quantity']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AX'), value=float(bill_dict['quantity']['AX']))
            price_cell.font = font_trebuchet_ms
            price_cell.number_format = '#,##0.00'
            price_cell.alignment = Alignment(horizontal='center')

        if 'AY' in bill_dict['quantity']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AY'), value=float(bill_dict['quantity']['AY']))
            price_cell.font = font_trebuchet_ms

        if 'AW' in bill_dict['quantity']:
            format_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AW'), value=bill_dict['quantity']['AW'])
            format_cell.number_format = '#,##0.00'

        if 'AC' in bill_dict['unit_price']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AC'), value=float(bill_dict['unit_price']['AC']))
            price_cell.number_format = 'R$ #,##0.00000'
            price_cell.font = font_trebuchet_ms
            price_cell.alignment = Alignment(horizontal='right')

        if 'AB' in bill_dict['unit_price']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AB'), value=float(bill_dict['unit_price']['AB']))
            price_cell.number_format = 'R$ #,##0.00000'
            price_cell.font = font_trebuchet_ms
            price_cell.alignment = Alignment(horizontal='right')

        if 'Q' in bill_dict['prices']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'Q'), value=float(bill_dict['prices']['Q']))
            price_cell.number_format = 'R$ #,##0.00'
            price_cell.font = font_trebuchet_ms

        if 'Z' in bill_dict['prices']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'Z'), value=float(bill_dict['prices']['Z']))
            price_cell.number_format = 'R$ #,##0.00'
            price_cell.font = font_trebuchet_ms
            price_cell.alignment = Alignment(horizontal='center')

        if 'S' in bill_dict['prices']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'S'), value=float(bill_dict['prices']['S']))
            price_cell.number_format = 'R$ #,##0.00'
            price_cell.font = font_trebuchet_ms

        if 'U' in bill_dict['prices']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'U'), value=float(bill_dict['prices']['U']))
            price_cell.number_format = 'R$ #,##0.00'
            price_cell.font = font_trebuchet_ms
            price_cell.alignment = Alignment(horizontal='right')

        if 'AT' in bill_dict['kwh_consumed']:
            format_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AT'), value=float(bill_dict['kwh_consumed']['AT']))
            format_cell.number_format = '#,##0.00'

        if 'AS' in bill_dict['kwh_consumed']:
            sheet.cell(row=max_row, column=column_index_from_string(
                'AS'), value=float(bill_dict['kwh_consumed']['AS']))

        if 'AU' in bill_dict['kwh_consumed']:
            sheet.cell(row=max_row, column=column_index_from_string(
                'AU'), value=float(bill_dict['kwh_consumed']['AU']))
    else:
        if 'AO' in bill_dict['quantity']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AO'), value=float(bill_dict['quantity']['AO']))
            price_cell.font = font_trebuchet_ms
            price_cell.number_format = '#,##0.00'
            price_cell.alignment = Alignment(horizontal='center')

        if 'AX' in bill_dict['quantity']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AX'), value=float(bill_dict['quantity']['AX']))
            price_cell.font = font_trebuchet_ms
            price_cell.number_format = '#,##0.00'
            price_cell.alignment = Alignment(horizontal='center')

        if 'AC' in bill_dict['unit_price']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AC'), value=float(bill_dict['unit_price']['AC']))
            price_cell.number_format = 'R$ #,##0.00000'
            price_cell.font = font_trebuchet_ms
            price_cell.alignment = Alignment(horizontal='right')

        if 'AB' in bill_dict['unit_price']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AB'), value=float(bill_dict['unit_price']['AB']))
            price_cell.number_format = 'R$ #,##0.00000'
            price_cell.font = font_trebuchet_ms
            price_cell.alignment = Alignment(horizontal='right')

        if 'Z' in bill_dict['prices']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'Z'), value=float(bill_dict['prices']['Z']))
            price_cell.number_format = 'R$ #,##0.00'
            price_cell.font = font_trebuchet_ms
            price_cell.alignment = Alignment(horizontal='center')

        if 'U' in bill_dict['prices']:
            price_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'U'), value=float(bill_dict['prices']['U']))
            price_cell.number_format = 'R$ #,##0.00'
            price_cell.font = font_trebuchet_ms
            price_cell.alignment = Alignment(horizontal='right')

        if 'AT' in bill_dict['kwh_consumed']:
            format_cell = sheet.cell(row=max_row, column=column_index_from_string(
                'AT'), value=float(bill_dict['kwh_consumed']['AT']))
            format_cell.number_format = '#,##0.00'


def last_row_with_value(worksheet, uc, column_index):
    max_row = worksheet.max_row
    for row in range(max_row, 0, -1):
        cell_value = worksheet.cell(row=row, column=column_index).value
        if cell_value == uc:
            return row
    return None


def row_to_dict(worksheet, row_index):
    headers = [cell.value for cell in worksheet[1]]
    data = {}
    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row_index, column=col)
        data[header] = cell.value
    return data


def insert_sheet(sheet_path, bill_dict):
    workbook = openpyxl.load_workbook(sheet_path)
    sheet = workbook.active
    last_row_uc = find_last_row_value(sheet, bill_dict['uc'])

    value_column = column_index_from_string("O")
    max_row = sheet.max_row
    for linha in range(max_row, 0, -1):
        # Obter o valor da célula na coluna especificada
        valor = sheet.cell(row=linha, column=value_column).value

        # Verificar se a célula contém um valor não vazio
        if valor is not None:
            # Retornar a linha em que encontrou um valor não vazio
            max_row = linha + 1
            break
    organize_sheet_columns(sheet, max_row, bill_dict)
    duplicate_columns_value(sheet, last_row_uc, max_row)

    workbook.save(sheet_path)
    print(f"Valor inserido na planilha {sheet_path}")
