import datetime
import re

import openpyxl
from openpyxl.utils import column_index_from_string


def format_date(date):
    # Convert date from "DD/MM/YYYY" to "MARÇO/YYYY"
    date_obj = datetime.datetime.strptime(date, '%d/%m/%Y')
    months = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio',
        'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro',
        'Novembro', 'Dezembro'
    ]
    month_str = months[date_obj.month - 1]
    formatted_date = f"{month_str}/{date_obj.year}"
    return formatted_date


def pdf_text(text):
    inicio = text.find("R$")

    if inicio == -1:
        return None

    fim = text.find("\n", inicio)

    if fim != -1:
        return text[inicio:fim]
    else:
        return text[inicio:]


def get_info_rows(get_rows, text, get_type):
    text_init = "ENERGIA ATIVA FORNECIDA FP kWh"
    text_end = "ENERGIA GERAÇÃO - KWH RESERVADO"
    index_init = text.find(text_init)
    index_end = text.find(text_end) + len(text_end)
    text_part = text[index_init:index_end]
    rows = text_part.split('\n')

    row_list = []
    match = None
    for i, row in enumerate(rows, start=1):
        # Verificar se a linha está nas linhas desejadas
        if i in get_rows:
            # Usar expressão regular para encontrar o segundo número
            if get_type == "quantity":
                match = re.search(r'\s(\d{1,7},\d{2})\s', row)
            elif get_type == "unit_price":
                match = re.search(r'kWh\s+([\d.,]+)', row)
            elif get_type == "kwh_consumed":
                if i in (30, 32):
                    match = re.search(r'\b\d{1,3}(?:\.\d{3})*(?:,\d+)?\b', row)
                if i == 31:
                    match = re.search(
                        r'\b\d{1,}(?:\.\d{1,})?(?:,\d{1,2})\b', row)
                if match:
                    row_list.append(match.group())
                    continue
            else:
                if i == 17:
                    # Processar a linha "CONTRIB. ILUM. PÚBLICA - MUNICIPAL"
                    match = re.search(
                        r'(\d{1,3}(?:\.\d{3})*,\d{2})\s*(ITENS FINANCEIROS)', row)
                else:
                    match = re.search(r'(\d{1,3}(?:\.\d{3})*,\d{2})\s*$', row)
            if match:
                # Se encontrou um número, adicionar à lista de valores encontrados
                row_list.append(match.group(1))
    return row_list


def find_values(text, bill_group='A'):
    # Definindo padrões de expressões regulares para o valor e a data
    quantity = get_info_rows([1, 2, 3, 7, 9, 11], text, "quantity")
    unit_prices = get_info_rows([1, 3, 13, 15], text, "unit_price")
    prices = get_info_rows([4, 6, 16, 17, 18, 19], text, "prices")
    kwh_consumed = get_info_rows([30, 31, 32], text, "kwh_consumed")

    price_default = r'R\$.*?(\d{1,3}(?:\.\d{3})*,\d{2})'
    date_default = r'(\d{2}/\d{2}/\d{4})'
    interval_default = r'(\d{2}/\d{2}/\d{4})'
    uc_default = r'UC (\d+)'

    # Procurando pelo padrão do valor
    match_valor = re.search(price_default, text)
    match_date = re.findall(date_default, text)
    match_uc = re.search(uc_default, text)
    match_interval = re.findall(interval_default, text)

    if match_valor:
        price = match_valor.group(1)
    if match_date:
        date = match_date[-1]
        date = format_date(date)
    if match_interval:
        date_match_interval = match_interval[2:3]
    if match_uc:
        uc = match_uc.group(1)

    cycle = f"{date_match_interval[0]} a {date_match_interval[1]}"
    unit_prices = [numero.replace('.', '').replace(',', '.')
                   for numero in unit_prices]
    unit_prices = [float(numero) for numero in unit_prices]

    prices = [numero.replace('.', '').replace(',', '.') for numero in prices]
    prices = [float(numero) for numero in prices]
    if bill_group == "A":
        unit_prices = [
            unit_prices[0]+unit_prices[2],
            unit_prices[1]+unit_prices[3],
        ]
        prices = [
            prices[0],
            prices[1]+prices[-1],
            prices[2],
            prices[3]
        ]
    else:
        unit_prices = unit_prices
        prices = prices
    bill_dict = {
        'price': price,
        'date': date,
        'cycle': cycle,
        'uc': uc,
        'quantity': quantity,
        'unit_price': unit_prices,
        'prices': prices,
        'kwh_consumed': kwh_consumed
    }
    return bill_dict


def insert_sheet(sheet_path, bill_dict):
    workbook = openpyxl.load_workbook(sheet_path)
    sheet = workbook.active

    value_column = column_index_from_string("O")
    max_row = sheet.max_row
    for linha in range(max_row, 0, -1):
        # Obter o valor da célula na coluna especificada
        valor = sheet.cell(row=linha, column=value_column).value

        # Verificar se a célula contém um valor não vazio
        if valor is not None:
            # Retornar a linha em que encontrou um valor não vazio
            max_row = linha + 1
            break

    sheet.cell(row=max_row, column=column_index_from_string(
        'M'), value=bill_dict['date'])
    sheet.cell(row=max_row, column=column_index_from_string(
        'N'), value=bill_dict['cycle'])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'O'), value=int(bill_dict['price']))
    price_cell.number_format = 'R$ #,##0.00'
    price_cell = sheet.cell(
        row=max_row, column=column_index_from_string('B'), value=bill_dict['uc'])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AO'), value=bill_dict['quantity'][0])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AP'), value=bill_dict['quantity'][1])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AN'), value=bill_dict['quantity'][2])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AX'), value=bill_dict['quantity'][3])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AY'), value=bill_dict['quantity'][4])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AW'), value=bill_dict['quantity'][5])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AC'), value=bill_dict['unit_price'][0])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AB'), value=bill_dict['unit_price'][1])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'Q'), value=bill_dict['prices'][0])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'Z'), value=bill_dict['prices'][1])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'S'), value=bill_dict['prices'][2])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'U'), value=bill_dict['prices'][3])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AT'), value=bill_dict['kwh_consumed'][0])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AS'), value=bill_dict['kwh_consumed'][1])
    price_cell = sheet.cell(row=max_row, column=column_index_from_string(
        'AU'), value=bill_dict['kwh_consumed'][2])

    workbook.save(sheet_path)
    print(f"Valor inserido na planilha {sheet_path}")
