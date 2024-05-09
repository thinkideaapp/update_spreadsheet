from datetime import datetime
import csv

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import NamedStyle


def convert_date(date_str, reverse=False):

    if reverse:
        month_mapping = {
            "January": "Janeiro",
            "February": "Fevereiro",
            "March": "Março",
            "April": "Abril",
            "May": "Maio",
            "June": "Junho",
            "July": "Julho",
            "August": "Agosto",
            "September": "Setembro",
            "October": "Outubro",
            "November": "Novembro",
            "December": "Dezembro"
        }
        # Separando o ano e o mês da string da data
        ano, mes_numero = date_str.split('-')

        # Convertendo o número do mês para o nome do mês
        mes_nome = list(month_mapping.keys())[int(mes_numero) - 1]

        # Montando a string da data no formato desejado
        data_formatada = f"{month_mapping[mes_nome]}/{ano}"

        return data_formatada
    else:
        month_mapping = {
            'Janeiro': 'January',
            'Fevereiro': 'February',
            'Março': 'March',
            'Abril': 'April',
            'Maio': 'May',
            'Junho': 'June',
            'Julho': 'July',
            'Agosto': 'August',
            'Setembro': 'September',
            'Outubro': 'October',
            'Novembro': 'November',
            'Dezembro': 'December'
        }
        # Converte o nome do mês para inglês usando o dicionário de mapeamento
        for pt_month, en_month in month_mapping.items():
            date_str = date_str.replace(pt_month, en_month)
        # Converte a string da data para um objeto datetime
        date_obj = datetime.strptime(date_str, '%B/%Y')
        # Formata a data para o formato desejado 'YYYY-MM'
        return date_obj.strftime('%Y-%m')


def read_last_row(planilha, uc):
    # Inicializar um dicionário vazio
    dados_dict = {}

    # Abrir o arquivo CSV
    with open(planilha, newline='') as csvfile:
        # Ler o arquivo CSV usando o delimitador ';'
        reader = csv.reader(csvfile, delimiter=';')

        # Iterar sobre as linhas do arquivo CSV
        for linha in reader:
            # Ajustar o índice de acordo com a posição da coluna
            valor_b = str(linha[2]).split('- ')[-1]

            # Verificar se o valor na coluna B é igual a 109228
            if valor_b == uc:
                # Atualizar o dicionário com os valores da última linha
                for idx, valor in enumerate(linha):
                    # Usar str(idx + 1) como chave para o dicionário
                    dados_dict[str(idx + 1)] = valor

                print(f'Valor encontrado na linha: {linha}')

    return dados_dict


def find_last_row_value(ws, uc1, date1=None, reverse=False):
    # Inicia a busca na linha da última linha preenchida
    num_rows = ws.max_row
    while num_rows > 0:
        if ws.cell(row=num_rows, column=2).value == int(uc1):
            try:
                date2 = convert_date(
                    str(ws.cell(row=num_rows, column=13).value), reverse)
            except:
                date2 = str(ws.cell(row=num_rows, column=13).value)
            if date1:
                if date2 == date1:
                    return num_rows
            else:
                return num_rows
        num_rows -= 1
    # Retorna None se não encontrar o valor
    return None


def get_xlsx_uc(planilha1, planilha2, bill_dict):
    # Lê a última linha da primeira planilha
    last_row_dict = read_last_row(
        planilha1, bill_dict['uc'])
    # Carrega a segunda planilha
    wb = load_workbook(planilha2)
    # Seleciona a primeira planilha
    ws = wb.active
    # Procura a linha com o valor desejado na coluna B
    last_row = find_last_row_value(
        ws, bill_dict['uc'], bill_dict['date'], True)
    if last_row and last_row_dict:
        # Se encontrar, insere os valores de G, H e I
        try:
            value_ad = round(float(last_row_dict.get('7', 0)), 2)
            value_ae = round(float(last_row_dict.get('8', 0)), 2)
            value_af = round(float(last_row_dict.get('9', 0)), 2)
        except:
            value_ad = last_row_dict.get('7', 0)
            value_ae = last_row_dict.get('8', 0)
            value_af = last_row_dict.get('9', 0)

        format_cell = ws.cell(row=last_row, column=column_index_from_string("AD"))
        format_cell.value = value_ad if value_ad != "None" else ""
        format_cell.number_format = '#,##0.00'
        format_cell = ws.cell(row=last_row, column=column_index_from_string("AE"))
        format_cell.value = value_ae if value_ae != "None" else ""
        format_cell.number_format = '#,##0.00'
        format_cell = ws.cell(row=last_row, column=column_index_from_string("AF"))
        format_cell.value = value_af if value_af != "None" else ""
        format_cell.number_format = '#,##0.00'
        # Salva as alterações na planilha
        wb.save(planilha2)
    else:
        print("Valor não encontrado na segunda planilha.")
